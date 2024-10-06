import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
import csv
import os

class XlsxToCsvConverter:
    def __init__(self, root):
        self.root = root
        self.root.title('Conversor de Relatórios Solidcon')
        
        self.lbl_xlsx = tk.Label(root, text='Arquivo .xlsx:')
        self.lbl_xlsx.grid(row=0, column=0, padx=10, pady=5)
        
        self.edt_xlsx_path = tk.Entry(root, width=50)
        self.edt_xlsx_path.grid(row=0, column=1, padx=10, pady=5)
        
        self.btn_select_xlsx = tk.Button(root, text='Selecionar Arquivo .xlsx', command=self.select_xlsx)
        self.btn_select_xlsx.grid(row=0, column=2, padx=10, pady=5)
        
        self.lbl_csv = tk.Label(root, text='Salvar Arquivo .csv Como:')
        self.lbl_csv.grid(row=1, column=0, padx=10, pady=5)
        
        self.edt_csv_path = tk.Entry(root, width=50)
        self.edt_csv_path.grid(row=1, column=1, padx=10, pady=5)
        
        self.btn_save_csv = tk.Button(root, text='Salvar Arquivo .csv', command=self.save_csv)
        self.btn_save_csv.grid(row=1, column=2, padx=10, pady=5)
        
        self.btn_convert = tk.Button(root, text='Converter', command=self.convert)
        self.btn_convert.grid(row=2, column=1, padx=10, pady=10)
        
        self.btn_cancel = tk.Button(root, text='Cancelar', command=root.quit)
        self.btn_cancel.grid(row=2, column=2, padx=10, pady=10)
    
    def select_xlsx(self):
        filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filepath:
            self.edt_xlsx_path.delete(0, tk.END)
            self.edt_xlsx_path.insert(0, filepath)
    
    def save_csv(self):
        filepath = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
        if filepath:
            self.edt_csv_path.delete(0, tk.END)
            self.edt_csv_path.insert(0, filepath)
    
    def convert(self):
        xlsx_file = self.edt_xlsx_path.get()
        csv_file = self.edt_csv_path.get()
        
        if not xlsx_file or not csv_file:
            messagebox.showerror("Erro", "Por favor, selecione os arquivos .xlsx e .csv.")
            return
        
        try:
            wb = load_workbook(xlsx_file)
            sheet = wb.active
            with open(csv_file, mode='w', newline='') as file:
                writer = csv.writer(file)
                for row in sheet.iter_rows(values_only=True):
                    writer.writerow([cell if cell is not None else '' for cell in row])
            
            os.startfile(os.path.dirname(csv_file))
            messagebox.showinfo("Sucesso", "Conversão concluída com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao converter o arquivo: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = XlsxToCsvConverter(root)
    root.mainloop()
