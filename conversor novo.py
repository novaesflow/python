import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl

class MainApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Remover Células Mescladas")
        
        self.create_widgets()
        
    def create_widgets(self):
        self.input_file_label = tk.Label(self.root, text="Arquivo Origem:")
        self.input_file_label.grid(row=0, column=0, padx=10, pady=10, sticky='e')
        
        self.input_file_entry = tk.Entry(self.root, width=50)
        self.input_file_entry.grid(row=0, column=1, padx=10, pady=10)
        
        self.input_file_button = tk.Button(self.root, text="Importar", command=self.import_file)
        self.input_file_button.grid(row=0, column=2, padx=10, pady=10)
        
        self.output_file_label = tk.Label(self.root, text="Arquivo Destino:")
        self.output_file_label.grid(row=1, column=0, padx=10, pady=10, sticky='e')
        
        self.output_file_entry = tk.Entry(self.root, width=50)
        self.output_file_entry.grid(row=1, column=1, padx=10, pady=10)
        
        self.output_file_button = tk.Button(self.root, text="Salvar em", command=self.export_file)
        self.output_file_button.grid(row=1, column=2, padx=10, pady=10)
        
        self.convert_button = tk.Button(self.root, text="Converter", command=self.convert_file)
        self.convert_button.grid(row=2, column=1, padx=10, pady=10, sticky='e')
        
        self.cancel_button = tk.Button(self.root, text="Cancelar", command=self.cancel)
        self.cancel_button.grid(row=2, column=2, padx=10, pady=10, sticky='w')
    
    def import_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls *.xlsx")])
        self.input_file_entry.delete(0, tk.END)
        self.input_file_entry.insert(0, file_path)
    
    def export_file(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        self.output_file_entry.delete(0, tk.END)
        self.output_file_entry.insert(0, file_path)
    
    def convert_file(self):
        input_path = self.input_file_entry.get()
        output_path = self.output_file_entry.get()
        if not input_path or not output_path:
            messagebox.showerror("Erro", "Os campos de caminho do arquivo não podem estar vazios.")
            return
        try:
            wb = openpyxl.load_workbook(input_path)
            ws = wb.active

            # Cria um novo workbook para os dados processados
            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active

            # Desmesclar todas as células
            for merge in list(ws.merged_cells.ranges):
                ws.unmerge_cells(str(merge))
            
            # Copia os valores para o novo workbook
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                new_row = []
                for cell in row:
                    new_row.append(cell.value)
                new_ws.append(new_row)
            
            new_wb.save(output_path)
            messagebox.showinfo("Sucesso", "Arquivo processado com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao processar arquivo: {e}")
    
    def cancel(self):
        self.root.quit()

if __name__ == "__main__":
    root = tk.Tk()
    app = MainApp(root)
    root.mainloop()
